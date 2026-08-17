"""Import the distributor's real Excel workbook into the agent's data files.

  'Adam Chart' sheet -> data/festivals.json   (the festival corpus)
  'BAKARA'     sheet -> data/company.json     (company memory, PII stripped)

Contact names, e-mail addresses, invoice numbers and fees are never imported.

By default the distribution company is anonymised: its name becomes a fictional
one and its catalogue titles are replaced by stable pseudonyms, so the published
repository does not expose a real company's slate or festival strategy. The
festival facts and the relationship structure (which festival, how many
screenings, which years, which awards) are preserved exactly.
Pass --real-company to keep the original names for internal use.

Usage:  python scripts/import_excel.py "/path/to/workbook.xlsx"
        python scripts/import_excel.py "/path/to/workbook.xlsx" --real-company
"""

from __future__ import annotations

import argparse
import collections
import difflib
import hashlib
import json
import random
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# Adam Chart column indices (the sheet has three unheaded leading columns).
COL = {
    "tier": 2, "category": 3, "focus": 4, "secondary_focus": 5, "name": 6,
    "city": 7, "country": 8, "festival_dates": 9, "deadline_month": 10,
    "submission_open": 11, "next_deadline": 12, "final_deadline": 13,
    "status": 14, "premiere": 15, "fest_month": 16, "price": 17,
    "waiver": 18, "previous_films": 19, "website": 20,
}

# BAKARA column indices.
BAK = {"category": 1, "month": 4, "year": 5, "festival": 6, "country": 7, "film": 9, "award": 13}

ACCEPTS = {
    "iff": ["feature_fiction", "feature_doc", "short_fiction", "short_doc", "animation"],
    "go2friend iff": ["feature_fiction", "feature_doc", "short_fiction", "short_doc"],
    "doc": ["feature_doc", "short_doc"],
    "short iff": ["short_fiction", "short_doc", "animation"],
    "fiction": ["feature_fiction", "short_fiction"],
}

THEME_MAP = {
    "general": "general", "children and youth": "youth", "human rights": "human_rights",
    "ethnographic": "ethnographic", "women": "women_filmmakers", "mediterranean": "mediterranean",
    "debut": "first_feature", "art": "art", "disability": "disability", "music": "music",
    "lgbtq": "lgbtq", "religious": "religion", "romance": "romance", "scienece": "science",
    "science": "science", "short category": "short_form", "none": None,
}

REGIONS = {
    "Western Europe": {
        "France", "Germany", "Spain", "Italy", "Netherlands", "Belgium", "Austria", "Switzerland",
        "UK", "United Kingdom", "Scotland", "Ireland", "Portugal", "Luxembourg", "Luxemburg",
        "Monaco", "Malta",
    },
    "Northern Europe": {"Sweden", "Norway", "Denmark", "Finland", "Iceland", "Estonia", "Latvia", "Lithuania"},
    "Eastern Europe": {
        "Poland", "Czech Republic", "Czechia", "Czech", "Slovakia", "Hungary", "Romania", "Bulgaria",
        "Serbia", "Croatia", "Slovenia", "Bosnia and Herzegovina", "Bosnia and Hertzegovina", "Bosnia",
        "Macedonia", "North Macedonia", "Albania", "Montenegro", "Kosovo", "Ukraine", "Russia", "Belarus",
        "Moldova", "Georgia", "Armenia", "Greece", "Cyprus", "Turkey",
    },
    "North America": {"USA", "US", "United States", "Canada"},
    "Latin America": {
        "Mexico", "Brazil", "Argentina", "Chile", "Colombia", "Peru", "Uruguay", "Ecuador",
        "Bolivia", "Costa Rica", "Cuba", "Panama", "Guatemala", "Venezuela", "Paraguay",
    },
    "Middle East": {
        "Israel", "Palestine", "Jordan", "Lebanon", "Egypt", "UAE", "Qatar", "Iran", "Iraq",
        "Saudi Arabia", "Morocco", "Tunisia", "Algeria",
    },
    "Central Asia": {"Kazakhstan", "Uzbekistan", "Kyrgyzstan", "Tajikistan", "Azerbaijan"},
    "Africa": {"South Africa", "Kenya", "Nigeria", "Ghana", "Uganda", "Senegal", "Ethiopia", "Zimbabwe", "Rwanda"},
    "South Asia": {"India", "Nepal", "Sri Lanka", "Pakistan", "Bangladesh"},
    "East Asia": {"Japan", "South Korea", "Korea", "China", "Taiwan", "Hong Kong", "Mongolia"},
    "Southeast Asia": {"Singapore", "Thailand", "Vietnam", "Indonesia", "Malaysia", "Philippines", "Cambodia"},
    "Oceania": {"Australia", "New Zealand"},
}

STRATEGIC_VALUE = {
    "A": "Top-tier launch platform: selection creates international sales leverage and press.",
    "B+": "Strong international festival: meaningful visibility and a real step up for the film's profile.",
    "B": "Solid international slot: builds the festival record and reaches a defined audience.",
    "C": "Niche or regional slot: valuable for audience fit, low prestige cost if it does not land.",
}

NOISE = {"", "none", "n/a", "#n/a", "--", "-", "?", "tbd", "no info"}

ANON_COMPANY = {
    "id": "meridian-films",
    "name": "Meridian Films",
    "country": "Israel",
    "profile": (
        "Israeli international distributor and world sales agent for Israeli and "
        "Jewish-interest cinema, handling documentary and fiction features plus shorts. "
        "Runs a large recurring circuit of Jewish film festivals, cultural institutions and "
        "embassies alongside international festival submissions."
    ),
}

# The real company's details live in an untracked local file so the published
# repository never names the distributor. Same shape as ANON_COMPANY.
REAL_COMPANY_PATH = ROOT / "data" / "real_company.json"


def load_real_company() -> dict:
    if not REAL_COMPANY_PATH.exists():
        raise SystemExit(
            f"--real-company needs {REAL_COMPANY_PATH}, an untracked file containing "
            '{"id": ..., "name": ..., "country": ..., "profile": ...}.'
        )
    return json.loads(REAL_COMPANY_PATH.read_text(encoding="utf-8"))

# Word pools for stable pseudonymous film titles.
TITLE_A = [
    "Salt", "Ash", "Border", "Winter", "Harvest", "Iron", "Paper", "River", "Desert", "Stone",
    "Amber", "Quiet", "Hollow", "Northern", "Second", "Broken", "Distant", "Open", "Last", "First",
    "Golden", "Silent", "Narrow", "Bright", "Bitter", "Long", "Small", "Deep", "Wild", "Old",
]
TITLE_B = [
    "Road", "Season", "Letters", "House", "Garden", "Field", "Window", "Return", "Crossing",
    "Wedding", "Archive", "Kitchen", "Chorus", "Passage", "Harbour", "Lesson", "Portrait",
    "Inheritance", "Testimony", "Threshold", "Migration", "Anthem", "Vigil", "Trade", "Shelter",
    "Language", "Distance", "Promise", "Silence", "Bureau",
]
TITLE_SOLO = [
    "Afterlight", "Ashfall", "Northbound", "Homeland", "Undertow", "Groundwork", "Nightshift",
    "Aftermath", "Watershed", "Crosswind", "Landfall", "Foreshore", "Wintering", "Kinship",
]


class Anonymiser:
    """Deterministic, collision-free pseudonyms for catalogue titles."""

    def __init__(self, enabled: bool) -> None:
        self.enabled = enabled
        self.mapping: dict[str, str] = {}
        self._used: set[str] = set()

    def title(self, original: str | None) -> str | None:
        if not self.enabled or not original:
            return original
        if original in self.mapping:
            return self.mapping[original]

        seed = int(hashlib.sha1(original.encode("utf-8")).hexdigest()[:12], 16)
        rng = random.Random(seed)
        for attempt in range(60):
            if attempt < 40:
                candidate = f"{rng.choice(TITLE_A)} {rng.choice(TITLE_B)}"
            else:
                candidate = rng.choice(TITLE_SOLO)
            if attempt > 20:
                candidate = f"The {candidate}"
            if candidate not in self._used:
                break
        else:
            candidate = f"Untitled {len(self._used) + 1}"

        self._used.add(candidate)
        self.mapping[original] = candidate
        return candidate


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_or_none(value) -> str | None:
    text = clean(value)
    return None if text.lower() in NOISE else text


def month_name(value) -> str | None:
    text = clean(value)
    if not text:
        return None
    try:
        number = int(float(text))
    except ValueError:
        return text if text in MONTHS else None
    return MONTHS[number - 1] if 1 <= number <= 12 else None


def iso_date(value) -> str | None:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = clean(value)
    return text if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text) else None


def slugify(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", normalized).strip("-").lower()
    return slug[:60] or "festival"


def normalize_name(name: str) -> str:
    text = unicodedata.normalize("NFKD", str(name or "")).encode("ascii", "ignore").decode().lower()
    text = re.sub(r"\(.*?\)", " ", text)
    text = re.sub(
        r"\b(international|film|festival|fest|filmfest|iff|the|of|for|and|de|du|des|la|le|les|cinema|cinematography|movie|movies)\b",
        " ", text,
    )
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def region_for(country: str | None) -> str | None:
    if not country:
        return None
    for region, members in REGIONS.items():
        if country in members:
            return region
    return None


def parse_premiere(raw: str | None) -> tuple[str, str | None]:
    """Return (level, territory). 'World - Turkey' = world premiere, Turkish territory."""

    if not raw:
        return "unknown", None
    text = raw.strip()
    lowered = text.lower()
    if lowered.startswith("no requirement"):
        return "none", None
    if lowered.startswith("no info"):
        return "unknown", None
    if lowered.startswith("world"):
        territory = re.sub(r"^world\s*-?\s*", "", text, flags=re.I).strip(" -")
        return "world", territory or None
    return "unknown", text


def parse_previous_films(raw: str | None) -> list[dict]:
    """Parse 'Real Estate (2024) $0, We Will Dance Again (2024) $0'."""

    if not raw or raw.strip().lower() in NOISE:
        return []
    films = []
    for match in re.finditer(r"([^,(]+?)\s*\((\d{4})\)", raw):
        title = match.group(1).strip(" ,;\"'")
        if title:
            films.append({"title": title, "year": int(match.group(2))})
    return films


# ---------------------------------------------------------------- festivals
def import_festivals(workbook, anonymiser: Anonymiser) -> list[dict]:
    sheet = workbook["Adam Chart"]
    rows = [
        row for row in sheet.iter_rows(min_row=2, values_only=True)
        if any(cell is not None and str(cell).strip() for cell in row)
    ]

    festivals: list[dict] = []
    used_ids: set[str] = set()
    skipped = 0

    for row in rows:
        def cell(key):
            index = COL[key]
            return row[index] if index < len(row) else None

        name = clean_or_none(cell("name"))
        if not name:
            skipped += 1
            continue

        slug = slugify(name)
        candidate, suffix = slug, 2
        while candidate in used_ids:
            candidate, suffix = f"{slug}-{suffix}", suffix + 1
        used_ids.add(candidate)

        category = (clean(cell("category")) or "IFF").strip()
        accepts = ACCEPTS.get(category.lower(), ACCEPTS["iff"])

        themes = []
        for key in ("focus", "secondary_focus"):
            mapped = THEME_MAP.get(clean(cell(key)).lower(), clean_or_none(cell(key)))
            if mapped and mapped not in themes:
                themes.append(mapped)
        if category.lower() == "doc" and "documentary" not in themes:
            themes.insert(0, "documentary")
        if category.lower() == "short iff" and "short_form" not in themes:
            themes.insert(0, "short_form")

        premiere_raw = clean_or_none(cell("premiere"))
        level, territory = parse_premiere(premiere_raw)
        tier = (clean(cell("tier")) or "C").upper()
        country = clean_or_none(cell("country"))

        festivals.append(
            {
                "id": candidate,
                "name": name,
                "city": clean_or_none(cell("city")),
                "country": country,
                "region": region_for(country),
                "tier": tier,
                "category": category,
                "accepts": accepts,
                "themes": themes,
                "month": month_name(cell("fest_month")),
                "festival_dates": clean_or_none(cell("festival_dates")),
                "typical_deadline_month": month_name(cell("deadline_month")),
                "submission_open": iso_date(cell("submission_open")),
                "next_deadline": iso_date(cell("next_deadline")),
                "final_deadline": iso_date(cell("final_deadline")),
                "status": clean_or_none(cell("status")),
                "premiere_requirement": level,
                "premiere_requirement_raw": premiere_raw,
                "premiere_territory": territory,
                "submission_fee": clean_or_none(cell("price")),
                "waiver": clean_or_none(cell("waiver")),
                "website": clean_or_none(cell("website")),
                "company_previous_films": [
                    {"title": anonymiser.title(entry["title"]), "year": entry["year"]}
                    for entry in parse_previous_films(clean_or_none(cell("previous_films")))
                ],
                "strategic_value": STRATEGIC_VALUE.get(tier, STRATEGIC_VALUE["C"]),
                # Filled by scripts/merge_enrichment.py — descriptive text only.
                "focus": None,
                "award_patterns": None,
                "notable_past_selections": [],
                "notes": None,
                "source": "Adam Chart (distributor workbook)",
            }
        )

    print(f"festivals: {len(festivals)} imported, {skipped} rows skipped (no name)")
    return festivals


# ------------------------------------------------------------ company memory
def import_company(workbook, festivals: list[dict], anonymiser: Anonymiser, company: dict) -> dict:
    sheet = workbook["BAKARA"]
    rows = [
        row for row in sheet.iter_rows(min_row=3, values_only=True)
        if any(cell is not None and str(cell).strip() for cell in row)
    ]

    lookup = {normalize_name(f["name"]): f["id"] for f in festivals}
    keys = list(lookup)
    resolved: dict[str, str | None] = {}

    def resolve(name: str) -> str | None:
        if name not in resolved:
            key = normalize_name(name)
            if key in lookup:
                resolved[name] = lookup[key]
            else:
                close = difflib.get_close_matches(key, keys, n=1, cutoff=0.9)
                resolved[name] = lookup[close[0]] if close else None
        return resolved[name]

    per_festival: dict[str, dict] = {}
    film_stats: dict[str, dict] = {}
    venue_counts: collections.Counter = collections.Counter()
    category_counts: collections.Counter = collections.Counter()
    country_counts: collections.Counter = collections.Counter()
    matched_rows = 0

    for row in rows:
        def cell(key):
            index = BAK[key]
            return row[index] if index < len(row) else None

        festival_name = clean_or_none(cell("festival"))
        film = anonymiser.title(clean_or_none(cell("film")))
        year_raw = clean(cell("year"))
        year = int(float(year_raw)) if re.fullmatch(r"\d{4}(\.0)?", year_raw) else None
        award = clean_or_none(cell("award"))
        category = clean_or_none(cell("category"))
        country = clean_or_none(cell("country"))

        if category:
            category_counts[category.strip(". ").title()] += 1
        if country:
            country_counts[country] += 1
        if festival_name:
            venue_counts[festival_name] += 1

        if film:
            stats = film_stats.setdefault(film, {"title": film, "screenings": 0, "first_year": year, "last_year": year, "awards": []})
            stats["screenings"] += 1
            if year:
                stats["first_year"] = min(stats["first_year"] or year, year)
                stats["last_year"] = max(stats["last_year"] or year, year)
            if award and award not in stats["awards"]:
                stats["awards"].append(award)

        if not festival_name:
            continue
        festival_id = resolve(festival_name)
        if not festival_id:
            continue

        matched_rows += 1
        record = per_festival.setdefault(
            festival_id,
            {
                "company_id": company["id"],
                "festival_id": festival_id,
                "festival_name": festival_name,
                "screenings": 0,
                "films": [],
                "years": [],
                "awards": [],
                "categories": [],
            },
        )
        record["screenings"] += 1
        if film and film not in record["films"]:
            record["films"].append(film)
        if year and year not in record["years"]:
            record["years"].append(year)
        if award:
            record["awards"].append({"film": film, "year": year, "award": award})
        if category and category not in record["categories"]:
            record["categories"].append(category)

    # Fold in the "Previous Films" column of Adam Chart.
    for festival in festivals:
        previous = festival.get("company_previous_films") or []
        if not previous:
            continue
        record = per_festival.setdefault(
            festival["id"],
            {
                "company_id": company["id"], "festival_id": festival["id"],
                "festival_name": festival["name"], "screenings": 0,
                "films": [], "years": [], "awards": [], "categories": [],
            },
        )
        for entry in previous:
            if entry["title"] not in record["films"]:
                record["films"].append(entry["title"])
            if entry["year"] not in record["years"]:
                record["years"].append(entry["year"])
        record["screenings"] = max(record["screenings"], len(previous))

    history = []
    for record in per_festival.values():
        record["years"].sort()
        record["films"] = record["films"][:12]
        record["result"] = "awarded" if record["awards"] else "selected"
        record["note"] = build_note(record)
        history.append(record)
    history.sort(key=lambda record: record["screenings"], reverse=True)

    films = sorted(film_stats.values(), key=lambda item: item["screenings"], reverse=True)

    company_record = {
        **company,
        "circuit_summary": {
            "screening_records": len(rows),
            "distinct_venues": len(venue_counts),
            "catalogue_titles": len(film_stats),
            "years_covered": "2008-2027",
            "top_venue_types": category_counts.most_common(8),
            "top_countries": country_counts.most_common(10),
            "top_venues": [] if anonymiser.enabled else venue_counts.most_common(15),
        },
        "films": films[:150],
    }

    print(
        f"company memory: {len(history)} festivals with history "
        f"({matched_rows} of {len(rows)} screening rows matched the festival list)"
    )
    return {"company": company_record, "history": history}


def build_note(record: dict) -> str:
    parts = []
    years = record["years"]
    if years:
        span = f"{years[0]}" if len(years) == 1 else f"{years[0]}-{years[-1]}"
        parts.append(f"{record['screenings']} screening(s) between {span}")
    if record["films"]:
        parts.append("titles: " + ", ".join(record["films"][:5]))
    if record["awards"]:
        awards = "; ".join(
            f"{award['award']} ({award['film']}, {award['year']})" for award in record["awards"][:3]
        )
        parts.append(f"awards: {awards}")
    return ". ".join(parts) + "." if parts else "Prior relationship on record."


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument(
        "--real-company",
        action="store_true",
        help="keep the real company name and catalogue titles (internal use only)",
    )
    args = parser.parse_args()

    path = Path(args.workbook).expanduser()
    if not path.exists():
        raise SystemExit(f"workbook not found: {path}")

    anonymiser = Anonymiser(enabled=not args.real_company)
    company = load_real_company() if args.real_company else ANON_COMPANY
    print(
        f"company: {company['name']} "
        f"({'real data' if args.real_company else 'anonymised'})"
    )

    workbook = openpyxl.load_workbook(path, data_only=True)
    festivals = import_festivals(workbook, anonymiser)
    company_data = import_company(workbook, festivals, anonymiser, company)

    data_dir = ROOT / "data"
    data_dir.mkdir(exist_ok=True)
    (data_dir / "festivals.json").write_text(
        json.dumps(festivals, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (data_dir / "company.json").write_text(
        json.dumps(company_data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"wrote {data_dir/'festivals.json'} and {data_dir/'company.json'}")

    if anonymiser.enabled and anonymiser.mapping:
        # Kept out of version control: the only link back to the real catalogue.
        mapping_path = data_dir / "anonymisation_map.json"
        mapping_path.write_text(
            json.dumps(anonymiser.mapping, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
        print(f"wrote {mapping_path} ({len(anonymiser.mapping)} titles pseudonymised, gitignored)")


if __name__ == "__main__":
    main()
